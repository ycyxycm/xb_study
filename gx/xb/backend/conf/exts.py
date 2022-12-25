# MALE_STYLECODE_LIST=[]
# FAMALE_STYLECODE_LIST=["男女宽松T", "男女中长圆领复合", "男女中长圆领毛圈","男女中长连帽复合","男女中长连帽毛圈","男女复合开衫","男女毛圈开衫"]
# CHILD_STYLECODE_LIST=["童纯棉T", "童连帽复合", "童连帽毛圈", "童圆领复合", "童圆领毛圈","童插肩防晒衣","童毛圈长裤","童复合长裤"]
import calendar
from datetime import datetime
from datetime import timedelta
from datetime import timezone
# smtplib 用于邮件的发信动作
import smtplib
# email 用于构建邮件内容
from email.mime.text import MIMEText
# 构建邮件头
from email.header import Header
import openpyxl
import random


from conf.config import *

#获取请求方ip
def get_request_ip(request):
    if 'HTTP_X_FORWARDED_FOR' in request.META:
        ip = request.META.get('HTTP_X_FORWARDED_FOR')
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

#SimpleCookie转字符串cookies
def cookies_simple(cookies_simple):
    print(type(cookies_simple))
    # if type(cookies_simple) == requests.cookies.
    cook=""
    for i in cookies_simple:
        cook=cook+';'+i+'='+cookies_simple[i].value
    return cook[1:]

#生成指定长度验证码
def generate_code(code_len=6):
    """
    生成指定长度验证码
    :param code_len: 验证码的长度(默认4个字符)
    :return: 由大小写英文字母以及数字构成的随机验证码
    """
    all_chars = '0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
    last_pos = len(all_chars) - 1
    code = ''
    for _ in range(code_len):
        index = random.randint(0, last_pos)
        code += all_chars[index]
    return code

#发送邮件
def send_email(addressee_email,vcode):
    from_addr=EMAIL
    password=EMAIL_IMAPSMTP
    #发件服务器
    smtp_server="smtp.163.com"

    # 邮箱正文内容，第一个参数为内容，第二个参数为格式(plain 为纯文本)，第三个参数为编码
    msg = MIMEText('验证码:'+vcode, 'plain', 'utf-8')
    # 邮件头信息
    msg['From'] = Header('小布system')  # 发送者
    msg['To'] = Header('员工')  # 接收者
    subject = '小布System验证邮件'
    msg['Subject'] = Header(subject, 'utf-8')  # 邮件主题

    try:
        smtpobj = smtplib.SMTP_SSL(smtp_server)
        # 建立连接--qq邮箱服务和端口号（可百度查询）
        smtpobj.connect(smtp_server, 465)
        # 登录--发送者账号和口令
        smtpobj.login(from_addr, password)
        # 发送邮件
        smtpobj.sendmail(from_addr, addressee_email, msg.as_string())
        print("邮件发送成功")
        return True
    except smtplib.SMTPException:
        print("无法发送邮件")
        return False
    finally:
        # 关闭服务器
        smtpobj.quit()

#小布判断性别
def get_now_sex(stylecode):#0=女 男=1 童=2
    male_stylecode=['男女JDS牛仔',"男女假两件牛仔","男女加绒假两件牛仔","男女破洞牛仔","男女条纹T","男女通穿宽松版短袖"]
    if stylecode in male_stylecode:
        return 1
    elif stylecode[:2]=="男女" or stylecode[:2]=="女通" or stylecode[0]=="女":
        return 0
    elif stylecode[0]=="童":
        return 2
    elif stylecode[0]=="男":
        return 1

#获取当前月份
def get_now_yearsmonth():
    SHA_TZ = timezone(
    timedelta(hours=8),
    name='Asia/Shanghai',
    )
    # 协调世界时
    utc_now = datetime.utcnow().replace(tzinfo=timezone.utc)
    # 北京时间
    beijing_now = utc_now.astimezone(SHA_TZ)
    now = datetime.now().date()
    #当前时间
    now_time=beijing_now.strftime("%Y-%m")
    now_years=beijing_now.strftime("%Y")
    now_month=beijing_now.strftime("%m")
    now_day=beijing_now.strftime("%d")

    this_month_start = datetime(now.year, now.month, 1).strftime('%Y-%m-%d')
    this_month_end = datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1]).strftime('%Y-%m-%d')

    return {"start_date": this_month_start, "end_date": this_month_end}
    # return {"now_date":now_time,"y":now_years,'m':now_month}

#获取当前日期
def get_now_date():
    SHA_TZ = timezone(
        timedelta(hours=8),
        name='Asia/Shanghai',
    )
    # 协调世界时
    utc_now = datetime.utcnow().replace(tzinfo=timezone.utc)
    # 北京时间
    beijing_now = utc_now.astimezone(SHA_TZ)
    # 当前时间
    now_time = beijing_now.strftime("%Y-%m-%d")
    return now_time


#openxl写入excel
def save_xlsx(_data, file_path):
    wb = openpyxl.Workbook()
    sd = wb.active
    for i in _data:
        sd.append(i)
    wb.save(file_path)

#读取xlsx文件
def read_xlsx(path, sheet_name=None, header_line=1):
    wb = openpyxl.load_workbook(path)
    if sheet_name:
        sd = wb[sheet_name]
    else:
        sd = wb.active
    data = [list(item) for item in sd.values]
    data = data[header_line:]
    return data

#读取xlsx的所有sheet页 分析多维度数据
def read_xlsx_dwd(path,header_line=1):
    wb = openpyxl.load_workbook(path)
    sheet_list=wb.sheetnames
    data_count=[]
    for i in sheet_list:
        if i=="查询条件":
            continue
        sd=wb[i]
        data=[list(item) for item in sd.values]
        data = data[header_line:]
        data_dict={'shop':i}
        for j in data:
            if j[1] is None:
                continue
            elif len(j)<5:
                val=0
            elif j[4] is None:
                val=0
            else:
                val=round(float(j[4].replace(",","")),2)
            name = j[1].replace("(+)", "").replace("(-)", "")
            data_dict[name]=val
        data_count.append(data_dict)
    return data_count


#颜色打印
def print_cyan(str_info):
    return print(f"\033[1;36m{str_info}\033[0m")

#除法
def chufa(a,b,default_val=0):
    try:
        re_number=a/b
    except ZeroDivisionError:
        re_number=default_val
    return re_number


# read_xlsx_dwd(path=r"C:\Users\hwj\Desktop\gx\xb\backend\media\PDD_dwd_files\汇总表-按店铺拆分_2022-11-20.xlsx")