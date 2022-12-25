# MALE_STYLECODE_LIST=[]
# FAMALE_STYLECODE_LIST=["男女宽松T", "男女中长圆领复合", "男女中长圆领毛圈","男女中长连帽复合","男女中长连帽毛圈","男女复合开衫","男女毛圈开衫"]
# CHILD_STYLECODE_LIST=["童纯棉T", "童连帽复合", "童连帽毛圈", "童圆领复合", "童圆领毛圈","童插肩防晒衣","童毛圈长裤","童复合长裤"]
from copy import deepcopy
import warnings,os

import openpyxl

BASE_DIR=os.getcwd()

def join_path(path):
    return os.path.join(BASE_DIR,path)

warnings.filterwarnings("ignore")

def get_now_sex(stylecode):#0=女 男=1 童=2
    male_stylecode=['男女JDS牛仔',"男女假两件牛仔","男女加绒假两件牛仔","男女破洞牛仔","男女条纹T","男女通穿宽松版短袖"]
    if stylecode in male_stylecode:
        return 1
    elif stylecode[:2]=="男女" or stylecode[:2]=="女通" or stylecode[0]=="女":
        return 0
    elif stylecode[0]=="童":
        return 2
    elif stylecode[0]=="男":
        return 1

#读取openpyxl
def excel_to_dict(active):
    max_column = active.max_column
    max_row = active.max_row
    excel_head = {}
    excel_head_list = []
    data_list = []
    for index, row in enumerate(active.rows):
        if index != 0:
            new_excel_head = deepcopy(excel_head)
        for d_index, j in enumerate(row):
            if index == 0:
                excel_head[j.value] = ''
                excel_head_list.append(j.value)
            else:
                new_excel_head[excel_head_list[d_index]] = j.value
            # print(j, j.value)
        if index:
            data_list.append(new_excel_head)
        # if index == 2:
    print(len(excel_head_list))
    return data_list

#openxl写入excel
def save_file(_data, file_path):
    wb = openpyxl.Workbook()
    sd = wb.active
    for i in _data:
        sd.append(i)
    wb.save(file_path)

#读取xlsx
def read_xlsx(path, sheet_name, header_line):
    wb = openpyxl.load_workbook(path)
    if sheet_name:
        sd = wb[sheet_name]
    else:
        sd = wb.active
    data = [list(item) for item in sd.values]
    data = data[header_line:]
    return data
# b="男女中长连帽复合"
# a=get_now_sex(stylecode=b)
# if a==0:
#     print("女")
# elif a==1:
#     print("男")
# elif a==2:
#     print("童")
